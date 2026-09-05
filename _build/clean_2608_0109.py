# -*- coding: utf-8 -*-
"""CLEAN bo alert 26/08 - 01/09/2026 -> Data\score_clean_2608_0109.csv

NGUON : 35 file CSV trong Downloads\data_2608_0109, moi file = output tho 1 kich ban.
        Ten file = TEN TOPIC -> map sang ten KB qua sheet FINAL cua file Excel kich ban.

XU LY:
 1. Ten KB       : map tu 'Ten topics' -> 'Ten Kich ban' (sheet FINAL cua file kich ban)
 2. Level + diem : lay tu BANG DIEM 64 KB — dung nguon voi nap_ngay_moi.py
 3. ngay_alert   : tach tu hau to _backupDDMM cua ma_khach_hang
                   (da kiem chung: khop 100% voi ngay_neo)
 4. object_value : ma_khach_hang sau khi cat hau to
 5. SCORE NEW    : tinh bang cong thuc PP-D 2 tang (r = k = 0.5),
                   giong het build_real_lab.py
 6. Level dau ra : cat theo nguong 20 / 60 / 86
"""
import os, csv, io, re, json, unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from openpyxl import load_workbook

SRC = r'C:\Users\daicd\Downloads\data_2608_0109'
XLS = r'D:\Project F2DR\File lưu trữ thông tin all kịch bản\File điền thông tin kịch bản_hành vi.xlsx'
XL_DIEM = r'D:\Project F2DR\File lưu trữ thông tin all kịch bản\20260625_scoring ver 3.xlsx'
SHEET_DIEM = 'Bảng điểm 64 KB'
DATA = r'D:\Project F2DR\File lưu trữ thông tin all kịch bản\HTML Scoring\Data'
OUT = os.path.join(DATA, 'score_clean_2608_0109.csv')
LOG = os.path.join(DATA, 'log_clean_2608_0109.txt')

CAP = {'Low': 25, 'Medium': 50, 'High': 80, 'Very High': 100}
R0 = K0 = 0.5
NGUONG = [(0, 20, 'Low'), (20, 60, 'Medium'), (60, 86, 'High'), (86, 1e9, 'Very High')]

COT_RA = ['AlertID', 'CATEGORY', 'created_date', 'requestor', 'updated_date',
          'reject_reason', 'usecase_name', 'status', 'object_key', 'object_value',
          'last_risk_score', 'Level', 'Score_KB', 'viettel_bank_code',
          'PARTITION_DATE', 'ngay_alert', 'gio_alert', 'usecase_clean',
          'request_id', 'risk_level_kb', 'chi_tiet_alert']
BO_QUA = {'ma_khach_hang', 'request_id', 'thoi_gian_gd', 'ngay_neo', 'viettel_bank_code'}
RX_HAUTO = re.compile(r'^(.*?)_backup(\d{2})(\d{2})$')

o = io.StringIO()
def P(*a):
    print(*a, file=o)
    try: print(*a)
    except UnicodeEncodeError:
        print(' '.join(str(x) for x in a).encode('ascii', 'replace').decode())
def f(x): return format(int(x), ',')

def norm(s):
    s = re.sub(r'^\[SCORE\]_', '', str(s or ''))
    s = re.sub(r'^\[NEW\]_', '', s)
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd').replace('Đ', 'D')
    return re.sub(r'[^a-z0-9]', '', s.lower())

def lv_tu_diem(s):
    for lo, hi, t in NGUONG:
        if lo <= s < hi: return t
    return 'Very High'

P('=' * 104)
P('CLEAN BO ALERT 26/08 - 01/09/2026')
P('Chay luc: %s' % datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
P('=' * 104)

# ═══ 1a. Map TEN TOPIC -> TEN KICH BAN  (sheet FINAL) ═══
wb = load_workbook(XLS, read_only=True, data_only=True)
rows = list(wb['FINAL'].iter_rows(values_only=True))
hdr = [str(x or '').strip() for x in rows[0]]
ix = {h: i for i, h in enumerate(hdr) if h}
map_topic = {}
for r in rows[1:]:
    kb = str(r[ix['Tên Kịch bản']] or '').strip()
    if not kb.startswith('[NEW]'): continue
    tp = str(r[ix['Tên topics']] or '').strip()
    map_topic[norm(tp) if tp else norm(kb)] = dict(
        kb=kb,
        score_name=str(r[ix['Tên KB Score']] or '').strip(),
        nhom=str(r[ix['Nhóm kịch bản']] or '').strip())
P('Sheet FINAL: %d KB (map ten topic -> ten kich ban)' % len(map_topic))

# ═══ 1b. Lay Level + Score_KB tu BANG DIEM  (cung nguon voi nap_ngay_moi.py) ═══
ws_d = load_workbook(XL_DIEM, data_only=True)[SHEET_DIEM]
LUT = {}
for r in ws_d.iter_rows(min_row=2, values_only=True):
    ten, lv, di = r[2], r[3], r[9]
    if ten and lv and di is not None:
        LUT[norm(ten)] = (str(lv).strip().title(), int(di))
P('Bang diem "%s": %d khoa' % (SHEET_DIEM, len(LUT)))
P('')

# ═══ 2. Doc 35 file ═══
files = sorted(x for x in os.listdir(SRC) if x.lower().endswith('.csv'))
P('So file nguon: %d' % len(files))

recs = []
canh_bao = []
kb_info = {}
thong_ke = []

for fn in files:
    topic = fn[:-4]
    m = map_topic.get(norm(topic))
    if not m:
        canh_bao.append('KHONG MAP DUOC topic: %s' % topic); continue
    kb = m['kb']
    if norm(kb) not in LUT:
        canh_bao.append('KB CHUA CO TRONG BANG DIEM: %s' % kb); continue
    lv, sc = LUT[norm(kb)]
    kb_info[kb] = dict(sc=float(sc), cap=CAP[lv], lv=lv,
                       nhom=m['nhom'], score_name=m['score_name'])

    n = 0
    with open(os.path.join(SRC, fn), encoding='utf-8-sig', newline='') as fh:
        for row in csv.DictReader(fh):
            mk = (row.get('ma_khach_hang') or '').strip()
            if not mk: continue
            mm = RX_HAUTO.match(mk)
            if mm:
                obj = mm.group(1); ngay = '2026-%s-%s' % (mm.group(3), mm.group(2))
            else:
                obj = mk
                ngay = (row.get('ngay_neo') or (row.get('thoi_gian_gd') or '')[:10]).strip()[:10]
                canh_bao.append('Khong co hau to backup: %s' % mk[:30])
            if not ngay: continue
            tgd = (row.get('thoi_gian_gd') or '').strip()
            gio = tgd.replace('T', ' ')[:19]
            ct = {k: v for k, v in row.items()
                  if k and k not in BO_QUA and v not in (None, '', 'null')}
            recs.append(dict(
                kb=kb, obj=obj, ngay=ngay, gio=gio,
                rid=(row.get('request_id') or '').strip(),
                vbc=(row.get('viettel_bank_code') or '').strip(),
                ct=json.dumps(ct, ensure_ascii=False) if ct else ''))
            n += 1
    thong_ke.append((kb, m['nhom'], lv, float(sc), n))

P('So ban ghi doc duoc: %s' % f(len(recs)))
P('')

# ═══ 3. Tinh SCORE NEW cho tung luot (obj, ngay) ═══
def eff(base, cap, n, r):
    return base if cap <= base else cap - (cap - base) * (r ** (n - 1))

luot = defaultdict(lambda: defaultdict(int))     # (obj,ngay) -> {kb: so alert}
for x in recs:
    luot[(x['obj'], x['ngay'])][x['kb']] += 1

diem = {}
for key, kbs in luot.items():
    es = [eff(kb_info[k]['sc'], kb_info[k]['cap'], n, R0) for k, n in kbs.items()]
    M = max(es); prod = 1.0; used = False
    for e in es:
        if not used and e == M: used = True; continue
        prod *= (1 - e / 100)
    diem[key] = min(100.0, max(0.0, M + (100 - M) * K0 * (1 - prod)))

P('=' * 104)
P('SCORE NEW  (cong thuc PP-D 2 tang, r = k = %.1f)' % R0)
P('=' * 104)
ds = sorted(diem.values())
def pct(p): return ds[min(len(ds) - 1, int(len(ds) * p / 100))]
P('So luot (KH x ngay): %s' % f(len(ds)))
P('p50 = %.1f   p75 = %.1f   p90 = %.1f   p99 = %.1f   max = %.1f'
  % (pct(50), pct(75), pct(90), pct(99), ds[-1]))
P('')
c = Counter(lv_tu_diem(v) for v in diem.values())
P('%-12s %10s %10s' % ('Muc', 'So luot', 'Ti trong'))
P('-' * 104)
for _, _, t in NGUONG:
    P('%-12s %10s %9.2f%%' % (t, f(c[t]), 100 * c[t] / len(ds)))
P('%-12s %10s %9.2f%%' % ('TONG', f(len(ds)), 100.0))
P('   (nguong tam 20 / 60 / 86 — se chot lai sau)')
P('')

# ═══ 4. Ghi file ═══
out = []
for x in recs:
    k = (x['obj'], x['ngay'])
    s = diem[k]
    info = kb_info[x['kb']]
    out.append({
        'AlertID': '%s_%s' % (x['ngay'].replace('-', ''), x['rid']) if x['rid'] else '',
        'CATEGORY': info['nhom'],
        'created_date': x['gio'], 'requestor': 'QTRR', 'updated_date': x['gio'],
        'reject_reason': '',
        'usecase_name': info['score_name'] or ('[SCORE]_' + x['kb']),
        'status': '0', 'object_key': 'userId', 'object_value': x['obj'],
        'last_risk_score': round(s, 6), 'Level': lv_tu_diem(s),
        'Score_KB': int(info['sc']), 'viettel_bank_code': x['vbc'],
        'PARTITION_DATE': x['ngay'].replace('-', ''), 'ngay_alert': x['ngay'],
        'gio_alert': x['gio'], 'usecase_clean': x['kb'], 'request_id': x['rid'],
        'risk_level_kb': info['lv'], 'chi_tiet_alert': x['ct']})

P('=' * 104)
P('THEO KICH BAN')
P('=' * 104)
P('%-56s %-14s %-11s %6s %8s' % ('Kich ban', 'Nhom', 'Risk Level', 'Score', 'Alert'))
P('-' * 104)
for a, b, l, s, n in sorted(thong_ke, key=lambda x: -x[4]):
    P('%-56s %-14s %-11s %6.0f %8s' % (a.replace('[NEW]_', '')[:56], b[:14], l, s, f(n)))
P('-' * 104)
P('%-56s %-14s %-11s %6s %8s' % ('TONG', '', '', '', f(sum(x[4] for x in thong_ke))))
P('')

P('=' * 104)
P('THEO NGAY')
P('=' * 104)
ng = Counter(x['ngay'] for x in recs)
kh_ng = defaultdict(set)
for x in recs: kh_ng[x['ngay']].add(x['obj'])
P('%-14s %10s %12s' % ('Ngay', 'So alert', 'KH rieng'))
P('-' * 104)
for k in sorted(ng):
    P('%-14s %10s %12s' % (k, f(ng[k]), f(len(kh_ng[k]))))
P('-' * 104)
P('%-14s %10s %12s' % ('TONG', f(len(recs)), f(len(set(x['obj'] for x in recs)))))
P('')
dd = Counter(len(x['obj']) for x in recs)
P('Do dai object_value: %s' % dict(dd))
P('   (20 = hash MSISDN chuan; 9-10 = ma kenh/merchant)')
P('')

if canh_bao:
    P('CANH BAO (%d):' % len(canh_bao))
    for k, v in Counter(canh_bao).most_common(12):
        P('   %-70s x%d' % (k[:70], v))
    P('')

with open(OUT, 'w', encoding='utf-8-sig', newline='') as fh:
    w = csv.DictWriter(fh, fieldnames=COT_RA)
    w.writeheader(); w.writerows(out)

P('DA GHI: %s' % os.path.basename(OUT))
P('   %s dong x %d cot, utf-8-sig, phan cach dau phay' % (f(len(out)), len(COT_RA)))
open(LOG, 'w', encoding='utf-8').write(o.getvalue())
print('\nLog: %s' % LOG)
