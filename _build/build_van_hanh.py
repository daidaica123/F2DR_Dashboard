# -*- coding: utf-8 -*-
"""
BUILD F2DR_Van_Hanh.html — cong cu XU LY ALERT DINH KY HANG TUAN.

Giao dien dung Y HET he mau / bo cuc cua F2DR_Real_Lab:
  - 3 muc dau (CONG THUC SCORE / NGUONG IMPACT / PHAN BO SCORE) giu nguyen, o dau file
  - phan van hanh (theo ngay, nhom nghiep vu, kich ban, KH phong alert) nam duoi

Chay (tu thu muc HTML Scoring):
    py -3.10 _build\\build_van_hanh.py
    py -3.10 _build\\build_van_hanh.py --csv "Data\\score_clean_2608_0109.csv"

Nguon so KB moi nhom: sheet FINAL cot "Cau hinh len F2DR" co chu 'ok'.
"""
import pandas as pd, numpy as np, json, os, sys, argparse, re, unicodedata
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
PARENT = os.path.dirname(HERE)
OUT = os.path.join(PARENT, "F2DR_Van_Hanh.html")
TPL = os.path.join(HERE, "_van_hanh_template.html")


def _tim(*ung_vien):
    """Tra ve duong dan dau tien co that (chay duoc ca o may lan tren Cloud)."""
    for p in ung_vien:
        if p and os.path.exists(p):
            return p
    return ung_vien[0]


# CSV mac dinh: thu ca thu muc "data" (ban deploy) lan "Data" (ban goc tren may)
DEF_CSV = _tim(os.path.join(PARENT, "data", "score_clean_2608_0109.csv"),
               os.path.join(PARENT, "Data", "score_clean_2608_0109.csv"))
# So KB moi nhom lay tu sheet FINAL. Tren Cloud khong co file Excel -> dung
# ban JSON chot san trong _build (sinh boi chinh script nay khi chay o may).
XLS = _tim(os.path.join(HERE, "nhom_kb.json"),
           os.path.join(PARENT, "File điền thông tin kịch bản_hành vi.xlsx"),
           r"D:\Project F2DR\File lưu trữ thông tin all kịch bản"
           r"\File điền thông tin kịch bản_hành vi.xlsx")

CAP = {"Low": 25, "Medium": 50, "High": 80, "Very High": 100}
NG_MAC_DINH = [25, 55, 80]           # vach khoi tao cua 3 thanh keo
R0 = K0 = 0.7                        # r, k mac dinh cua cong thuc PP-D

ap = argparse.ArgumentParser()
ap.add_argument("--csv", default=DEF_CSV)
ap.add_argument("--out", default=OUT)
ap.add_argument("--xls", default=XLS, help="File Excel co sheet FINAL")
ap.add_argument("--dotbien", type=float, default=2.0,
                help="Nguong danh dau dot bien: ngay cao nhat / trung binh")
ap.add_argument("--apk", type=float, default=3.0,
                help="Nguong danh dau alert/KH cao cua 1 kich ban")
ap.add_argument("--topkh", type=int, default=80, help="So KH phong alert dua vao bang")
ap.add_argument("--ghichep", default=os.path.join(HERE, "ghi_chep_dieu_tra.json"),
                help="File JSON ghi chep dieu tra (muc 7). Khong co thi de trong.")
a = ap.parse_args()

print("=" * 72); print("BUILD F2DR VAN HANH"); print("=" * 72)
for p in (a.csv, TPL):
    if not os.path.exists(p): sys.exit("THIEU FILE: " + p)
print("CSV :", os.path.basename(a.csv))

d = pd.read_csv(a.csv, encoding="utf-8-sig", dtype={"object_value": str})
need = ["AlertID", "object_value", "usecase_clean", "ngay_alert", "CATEGORY"]
miss = [c for c in need if c not in d.columns]
if miss: sys.exit("CSV thieu cot: " + ", ".join(miss))

if "risk_level_kb" in d.columns and d.risk_level_kb.notna().any():
    d["lvkb"] = d.risk_level_kb
elif "Level" in d.columns:
    d["lvkb"] = d.Level
else:
    sys.exit("CSV thieu cot Level cua kich ban (risk_level_kb hoac Level)")
bad = set(d.lvkb.dropna().unique()) - set(CAP)
if bad: sys.exit("Level la %s -> bo sung vao CAP." % bad)
if "Score_KB" not in d.columns: sys.exit("CSV thieu cot Score_KB")

d["ngay"] = d.ngay_alert.astype(str).str[:10]
days = sorted(d.ngay.unique())
dayi = {v: i for i, v in enumerate(days)}
ND = len(days)
print("Ngay : %d  (%s -> %s)" % (ND, days[0], days[-1]))
print("Alert: {:,}  |  KH: {:,}".format(len(d), d.object_value.nunique()))


# ═════════ SO KB DA CAU HINH MOI NHOM (sheet FINAL) ═════════
def norm(s):
    s = re.sub(r"^\[(SCORE|NEW)\]_", "", str(s or ""))
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s.replace("đ", "d").replace("Đ", "D").lower())


NHOM_TONG, KB2NHOM = {}, {}
JS_NHOM = os.path.join(HERE, "nhom_kb.json")
try:
    if a.xls.lower().endswith(".json"):
        # Ban chot san — dung tren Streamlit Cloud (khong co file Excel)
        j = json.load(open(a.xls, encoding="utf-8"))
        NHOM_TONG = {k: int(v) for k, v in j["nhom_tong"].items()}
        KB2NHOM = j["kb2nhom"]
        print("Nhom (nhom_kb.json): %d nhom, %d KB da cau hinh"
              % (len(NHOM_TONG), sum(NHOM_TONG.values())))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(a.xls, read_only=True, data_only=True)
        rows = list(wb["FINAL"].iter_rows(values_only=True))
        hdr = [str(x or "").strip() for x in rows[0]]
        iKB = hdr.index("Tên Kịch bản")
        iNH = hdr.index("Nhóm kịch bản")
        iCH = next(i for i, h in enumerate(hdr) if "cấu hình" in h.lower())
        cnt = defaultdict(int)
        for r in rows[1:]:
            kb = str(r[iKB] or "").strip()
            if not kb.startswith("[NEW]"): continue
            nh = str(r[iNH] or "").strip()
            if "ok" in str(r[iCH] or "").lower():
                cnt[nh] += 1
                KB2NHOM[norm(kb)] = nh
        NHOM_TONG = dict(cnt)
        wb.close()
        print("Nhom (sheet FINAL): %d nhom, %d KB da cau hinh"
              % (len(NHOM_TONG), sum(NHOM_TONG.values())))
        # Chot lai ra JSON de ban tren Cloud dung duoc ma khong can Excel
        json.dump({"nhom_tong": NHOM_TONG, "kb2nhom": KB2NHOM},
                  open(JS_NHOM, "w", encoding="utf-8"),
                  ensure_ascii=False, indent=1)
        print("            -> da ghi %s" % os.path.basename(JS_NHOM))
except Exception as e:
    print("!! Khong doc duoc danh sach nhom (%s) -> chi dem KB co alert" % e)

# ═════════ BANG KICH BAN ═════════
kb = (d.groupby("usecase_clean")
        .agg(lv=("lvkb", "first"), sc=("Score_KB", "first"),
             cat=("CATEGORY", "first"), alert=("AlertID", "size"),
             kh=("object_value", "nunique"))
        .reset_index().sort_values("alert", ascending=False))
kb["ten"] = kb.usecase_clean.str.replace(r"^\[NEW\]_", "", regex=True)
kb["cap"] = kb.lv.map(CAP)
kbi = {n: i for i, n in enumerate(kb.usecase_clean)}

mat = np.zeros((len(kb), ND), dtype=int)
for r in d.itertuples():
    mat[kbi[r.usecase_clean], dayi[r.ngay]] += 1
kh_kb_ngay = defaultdict(set)
for r in d.itertuples():
    kh_kb_ngay[(kbi[r.usecase_clean], dayi[r.ngay])].add(r.object_value)
matkh = np.zeros((len(kb), ND), dtype=int)
for (i, j), s in kh_kb_ngay.items(): matkh[i, j] = len(s)

KB = []
for i, r in enumerate(kb.itertuples()):
    v = mat[i]
    tb = v.mean()
    dinh = int(v.max()); ngay_dinh = days[int(v.argmax())]
    dot = float(dinh / tb) if tb > 0 else 0.0
    apk = float(r.alert / r.kh) if r.kh else 0.0
    # nhom nghiep vu cua kich ban — de giao dien ⑤ sap xep gom nhom lai voi nhau.
    # Khong tra duoc trong sheet FINAL thi lay tam CATEGORY cua chinh alert.
    nhom_kb = KB2NHOM.get(norm(r.ten), str(r.cat))
    KB.append({"ten": r.ten, "cat": str(r.cat), "nhom": nhom_kb,
               "lv": r.lv, "sc": int(r.sc),
               "cap": int(r.cap), "alert": int(r.alert), "kh": int(r.kh),
               "tb": round(float(tb), 1), "dinh": dinh, "ngayDinh": ngay_dinh,
               "dot": round(dot, 2), "apk": round(apk, 2),
               "v": [int(x) for x in v], "vkh": [int(x) for x in matkh[i]]})

# ═════════ THEO NGAY ═════════
NGAY = []
kh_ngay = d.groupby("ngay").object_value.apply(set).to_dict()
da_thay = set()
for i, g in enumerate(days):
    s = kh_ngay.get(g, set())
    moi = len(s - da_thay); da_thay |= s
    sub = d[d.ngay == g]
    NGAY.append({"ngay": g, "alert": int(len(sub)), "kh": int(len(s)),
                 "khMoi": int(moi), "kb": int(sub.usecase_clean.nunique())})

# ═════════ LUOT (KH x ngay) — de tinh SCORE NEW tren trinh duyet ═════════
det = (d.groupby(["object_value", "ngay", "usecase_clean"]).AlertID.size()
         .reset_index(name="f"))
det["ki"] = det.usecase_clean.map(kbi)
det = det.sort_values(["object_value", "ngay"])
pack = det.groupby(["object_value", "ngay"])[["ki", "f"]].apply(
    lambda x: [int(v) for p in zip(x.ki, x.f) for v in p])
LUOT_KH = [str(k[0]) for k in pack.index]
LUOT_NG = [dayi[k[1]] for k in pack.index]
PICKS = pack.tolist()
LUOT_AL = [sum(p[i + 1] for i in range(0, len(p), 2)) for p in PICKS]
LUOT_NKB = [len(p) // 2 for p in PICKS]
print("Luot : {:,}  (KH x ngay)".format(len(PICKS)))

# ═════════ TOP KH PHONG ALERT ═════════
# Moi KH dua ra chi tiet THEO TUNG NGAY (de loc ngay tren giao dien):
#   ngay -> [alert trong ngay, [ [kbIdx, so lan], ... ] ]
# Score tinh tren trinh duyet theo r,k dang keo, KHONG chot san o day.
kh_tong = d.groupby("object_value").agg(
    alert=("AlertID", "size"), nkb=("usecase_clean", "nunique"),
    nngay=("ngay", "nunique")).reset_index().sort_values("alert", ascending=False)
top = kh_tong.head(a.topkh)
tap_top = set(top.object_value)

# gom san: (KH, ngay) -> {kbIdx: so lan}
ct_kh = defaultdict(lambda: defaultdict(int))
for r in d.itertuples():
    if r.object_value in tap_top:
        ct_kh[(r.object_value, dayi[r.ngay])][kbi[r.usecase_clean]] += 1

TOPKH = []
for r in top.itertuples():
    ngay_ct = {}
    for j in range(ND):
        m = ct_kh.get((r.object_value, j))
        if m:
            ngay_ct[j] = sorted(([int(k), int(v)] for k, v in m.items()),
                                key=lambda x: -x[1])
    TOPKH.append({"kh": str(r.object_value), "alert": int(r.alert),
                  "nkb": int(r.nkb), "nngay": int(r.nngay),
                  "ng": ngay_ct})

# ═════════ TAI PHAM ═════════
c = kh_tong.nngay.value_counts().sort_index()
TAIPHAM = [{"n": int(k), "kh": int(v)} for k, v in c.items()]

# ═════════ NHOM NGHIEP VU — DU 7 NHOM, nhom khong co alert dien 0 ═════════
# Ngoai so lieu toan ky, xuat them so lieu TUNG NGAY de giao dien loc theo ngay
# ma khong phai build lai. Khach hang phai dem theo tap hop (khong cong ngang
# duoc qua ngay), nen luu san so KH rieng cua moi (nhom, ngay).
co_alert = defaultdict(lambda: {"alert": 0, "kh": set(), "kb": set()})
theo_ngay = defaultdict(lambda: {"alert": 0, "kh": set(), "kb": set()})
for r in d.itertuples():
    nh = KB2NHOM.get(norm(r.usecase_clean), str(r.CATEGORY))
    x = co_alert[nh]
    x["alert"] += 1; x["kh"].add(r.object_value); x["kb"].add(r.usecase_clean)
    y = theo_ngay[(nh, dayi[r.ngay])]
    y["alert"] += 1; y["kh"].add(r.object_value); y["kb"].add(r.usecase_clean)

ten_nhom = sorted(set(NHOM_TONG) | set(co_alert),
                  key=lambda n: -co_alert[n]["alert"] if n in co_alert else 1)
NHOM = []
for nh in ten_nhom:
    x = co_alert.get(nh)
    # ng: chi so ngay -> [alert, so KH rieng, so KB no] — bo qua ngay khong co alert
    ng = {}
    for j in range(ND):
        y = theo_ngay.get((nh, j))
        if y:
            ng[j] = [y["alert"], len(y["kh"]), len(y["kb"])]
    NHOM.append({"ten": nh,
                 "alert": x["alert"] if x else 0,
                 "kh": len(x["kh"]) if x else 0,
                 "kb": len(x["kb"]) if x else 0,
                 "kbTong": int(NHOM_TONG.get(nh, len(x["kb"]) if x else 0)),
                 "ng": ng})
NHOM.sort(key=lambda x: -x["alert"])

# Tong so KH rieng moi ngay (dong TONG cua bang ④ khi loc 1 ngay) — lay tu NGAY
# da tinh o tren, khong cong ngang duoc tu cac nhom.

# ═════════ DANH DAU KB BAT THUONG (khong con muc canh bao rieng) ═════════
for k in KB:
    k["flagDot"] = bool(k["dot"] >= a.dotbien and k["alert"] >= 20)
    k["flagDay"] = bool(k["apk"] >= a.apk and k["alert"] >= 20)
nguong_kh = max(10, int(np.percentile(kh_tong.alert, 99.5)))
for t in TOPKH:
    t["flag"] = bool(t["alert"] >= nguong_kh)
n_dot = sum(1 for k in KB if k["flagDot"])
n_day = sum(1 for k in KB if k["flagDay"])
n_kh = sum(1 for t in TOPKH if t["flag"])
print("Danh dau: %d KB dot bien | %d KB ban day | %d KH phong (>=%d alert)"
      % (n_dot, n_day, n_kh, nguong_kh))

tb_ngay = float(np.mean([x["alert"] for x in NGAY]))

# ═════════ GHI CHEP DIEU TRA (muc 7) ═════════
# File JSON: [{"ngay": "...", "tieude": "...", "noidung": "<html ngan>", "tag": [...]}]
GHICHEP = []
if os.path.exists(a.ghichep):
    try:
        GHICHEP = json.load(open(a.ghichep, encoding="utf-8"))
        print("Ghi chep: %d muc" % len(GHICHEP))
    except Exception as e:
        print("!! Doc ghi chep loi (%s) -> de trong" % e)
else:
    print("Ghi chep: chua co (%s)" % os.path.basename(a.ghichep))

DATA = {
    "days": days, "ngay": NGAY, "kb": KB, "nhom": NHOM,
    "picks": PICKS, "luotKh": LUOT_KH, "luotNg": LUOT_NG,
    "luotAl": LUOT_AL, "luotNkb": LUOT_NKB,
    "topkh": TOPKH, "taipham": TAIPHAM, "ghichep": GHICHEP,
    "tong": {"alert": int(len(d)), "kh": int(d.object_value.nunique()),
             "kb": int(d.usecase_clean.nunique()),
             "kbCauHinh": int(sum(NHOM_TONG.values())) if NHOM_TONG else int(len(kb)),
             "nd": ND, "luot": len(PICKS), "tbNgay": round(tb_ngay, 1),
             "d0": days[0], "d1": days[-1]},
    "nguong": NG_MAC_DINH, "cap": CAP, "r": R0, "k": K0,
    "cauhinh": {"dotbien": a.dotbien, "apk": a.apk, "nguongKh": nguong_kh},
    "file": os.path.basename(a.csv),
}

tpl = open(TPL, encoding="utf-8").read()
html = tpl.replace("__DATA__", json.dumps(DATA, ensure_ascii=False,
                                          separators=(",", ":")))
open(a.out, "w", encoding="utf-8").write(html)
print("-" * 72)
print("XONG: %s  (%s bytes)" % (os.path.basename(a.out),
                                format(os.path.getsize(a.out), ",")))
print("      %d ngay: %s -> %s" % (ND, days[0], days[-1]))
